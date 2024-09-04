import re
from mindformers import LlamaTokenizer
import json
import openpyxl
import os
import numpy
import argparse

LLAMA2_tokenizer = "./tokenizer.model"


def init_tokenizer(model_path=LLAMA2_tokenizer):
    tokenizer = LlamaTokenizer(model_path)
    return tokenizer


def get_text_token_num(tokenizer, text):
    tokens = tokenizer.tokenize(text)
    num_tokens = len(tokens)
    print("文本中包含的token数量：", num_tokens)
    return num_tokens


def gen_init_data(files):
    for res_file in files:
        result = {}
        with open(res_file) as f:
            data = json.loads(f.read())
        tokenizer_ = init_tokenizer()
        res_time_all = 0
        token_num_all = 0
        for i, item in enumerate(data):
            print(item)
            token_num = get_text_token_num(tokenizer_, item['resp_text'])
            print(f"Input : {item['input']}; generate tokenNum is {token_num}, resTime is {item['res_time']}")
            token_num_all = token_num_all + token_num
            res_time_all = res_time_all + item['res_time']
            result[i] = {"Input": item['input'], "token_num": token_num, "resTime": item['res_time']}
        # data_json = os.path.join(out_dir, f"{res_file.split('.json')[0]}_data.json")
        with open(res_file, "w+") as f:
            f.write(json.dumps(result))


def generate_excel(x_rate, token_num_list_, res_time_list_, count_all=False):
    if not os.path.exists(xlsx_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Serving_e2e"
    else:
        wb = openpyxl.load_workbook(filename=xlsx_file)
        ws = wb["Serving_e2e"]
    start_column = 1
    # 确认初始列
    while True:
        print(ws.cell(row=2, column=start_column).value)
        if not ws.cell(row=2, column=start_column).value:
            break
        start_column += 1
    print(f"blank column is {start_column}")

    ws.cell(row=2, column=start_column).value = "e2eTime"
    ws.cell(row=2, column=start_column + 1).value = "genTokenNum"
    ws.cell(row=2, column=start_column + 2).value = "latency"
    ws.merge_cells(start_row=1, end_row=1, start_column=start_column, end_column=start_column + 2)
    ws.cell(row=1, column=start_column).value = f"X = {x_rate}"

    for i, token_num in enumerate(token_num_list_):
        token_num = 1 if token_num == 0 else token_num
        ws.cell(row=3 + i, column=start_column).value = res_time_list_[i]
        ws.cell(row=3 + i, column=start_column + 1).value = token_num
        ws.cell(row=3 + i, column=start_column + 2).value = res_time_list_[i] / token_num
    if count_all:
        ws.cell(row=2, column=start_column + 4).value = "x"
        ws.cell(row=2, column=start_column + 5).value = "latencyAvg"
        count_column = count_row = 3
        while count_column <= start_column + 2:
            print(count_column)
            row_value = 3
            latencies = []
            while True:
                if ws.cell(row=row_value, column=count_column).value:
                    latencies.append(ws.cell(row=row_value, column=count_column).value)
                    row_value += 1
                else:
                    break
            print(f"latencyAvg is {numpy.mean(latencies)}")
            print(f"x is {ws.cell(row=1, column=count_column - 2).value}")
            while True:
                print("111", ws.cell(row=count_row, column=start_column + 4).value)
                ws.cell(row=count_row, column=start_column + 4).value = ws.cell(row=1,
                                                                                column=count_column - 2).value = ws.cell(
                    row=1, column=count_column - 2).value
                ws.cell(row=count_row, column=start_column + 5).value = numpy.mean(latencies)
                count_row += 1
                if not ws.cell(row=count_row, column=start_column + 4).value:
                    break
            count_column += 3
    wb.save(xlsx_file)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="count serving performance")
    parser.add_argument("-O", "--out_dir", help='out dir', required=True)
    args = parser.parse_args()
    xlsx_file = os.path.join(args.out_dir, 'result.xlsx')
    if os.path.exists(xlsx_file):
        os.remove(xlsx_file)
    x_file_range = [os.path.join(args.out_dir, file_) for file_ in os.listdir(args.out_dir)]
    x_file_range.sort()
    print(x_file_range)
    gen_init_data(x_file_range)
    for i, x_file in enumerate(x_file_range):
        if_count_all = True if i == len(x_file_range) - 1 else False
        with open(x_file) as f:
            data = json.loads(f.read())
        print(data)
        token_num_list = [data[key]["token_num"] for key in data]
        res_time_list = [data[key]["resTime"] for key in data]
        print(token_num_list)
        print(res_time_list)
        generate_excel(re.findall("result_(.*)_x", x_file)[0], token_num_list, res_time_list, if_count_all)
